# -*- coding: utf-8 -*-
"""
Created on Wed Dec 25 19:36:45 2019

@author: hattab
"""
import numpy as np
import cv2
import glob
import os
from joblib import Parallel, delayed
import multiprocessing
import openpyxl
import fonctions_yal
import fonctions
from sklearn import preprocessing
from sklearn import  preprocessing

def cetre_des_donnees(data):
    Mean = data.mean(axis=0)
    Mean = np.ravel(Mean)
    return Mean


def les_donnees_center(data):
    df = data - cetre_des_donnees(data)
    return df


def calcul_eigen_vectors(data, N_CP=0):
    print(data.shape)
    print('calculer center')
    cd = les_donnees_center(data)
    print('calculer cov')
    cov = np.cov(cd)
    print('calculer eigen')
    eigen_vals, eigenVectors = np.linalg.eig(cov)

    # trier les vecteurs properes  par raport le triage des valeurs propres
    indices = np.argsort(-eigen_vals)
    eigen_vals = eigen_vals[indices]
    eigenVectors = eigenVectors[indices]
    # calculer le nombre des composantes principales

    if N_CP == 0:
        a = eigen_vals[0]
        N_CP = 0
        somme_eig_vals = sum(eigen_vals)
        while a / somme_eig_vals < 0.9999:
            N_CP = N_CP + 1
            a = a + eigen_vals[N_CP]
        N_CP = N_CP + 1
    print(N_CP)

    N_CP_eigenVectors = np.zeros((N_CP, cd.shape[0]))
    N_CP_eigen_vals = np.zeros((N_CP, 1))
    for i in range(0, N_CP):
        N_CP_eigenVectors[i] = eigenVectors[i]
        N_CP_eigen_vals[i] = eigen_vals[i]
    print(N_CP_eigenVectors.shape)

    #N_CP_eigenVectors = np.dot(N_CP_eigenVectors, cd) / np.sqrt(N_CP_eigen_vals)  # formule de transition
    return N_CP_eigenVectors


def Reconnaisance(w1, w2, k_nn, label_train, label_test):
    e = 0
    book = openpyxl.load_workbook(filename='a.xlsx')
    sheet = book.worksheets[0]
    print('11111')

    for ind in range(np.shape(label_test)[0]):
        print('photo', ind + 1)


        index = np.linalg.norm( w1-w2[ind], axis=1).argsort()
        # afficher 10 vissages
        '''fig, ax = plt.subplots(2, 5, figsize=(10, 5))
        for i in range(0, 10):
            ax[i // 5, i % 5].imshow(data[index[i]].reshape(112, 92), cmap="gray")
        plt.show()'''

        if fonctions_yal.calcul_knn(ind, index, k_nn, label_train, label_test) == 1:
            e = e + 1
            print('nombre d′image de teste reconnues ', e)
            sheet.cell(row=ind + 6, column=(k_nn // 2) + 2).value = 1
        else:
            print('erreur ')
            sheet.cell(row=ind + 6, column=(k_nn // 2) + 2).value = 0
    book.save("a.xlsx")


if __name__ == '__main__':


    train_path = "yal_train2"  # lien du dossier (train)
    test_path = "yal_test2"  # lien du dossier (test)
    num_cores = multiprocessing.cpu_count()#nombre des coeurs cpu

    #data = fonctions_yal.read_images(train_path)
    data=np.load('yal_bA+000_histo_2_ltp_train.npy',allow_pickle=True)
    data=fonctions.conv_list_arr(data,fonctions.nomb_img( train_path))#convertir  histogrammes_test en 2 dimensions puisque chaque coeur calculer une matrice dépendante



    N_CP_eigenVectors = calcul_eigen_vectors(data.T)

    w1 = les_donnees_center(data).dot(N_CP_eigenVectors.T)
    test = np.load('yal_bA+000_histo_2_ltp_test.npy',allow_pickle=True)
    test=fonctions.conv_list_arr(test,fonctions.nomb_img( test_path))#convertir  histogrammes_test en 2 dimensions puisque chaque coeur calculer une matrice dépendante

    test_c=test-cetre_des_donnees(data)
    print(test_c[0])
    w2 = (N_CP_eigenVectors.dot((test_c).T)).T
    label_test = fonctions.labels(test_path)
    label_train = fonctions.labels(train_path)
    print(w1.shape)
    print(w2.shape)
    histogrammes_train=w1
    histogrammes_test=w2

    mix = np.concatenate((histogrammes_train, histogrammes_test), axis=0)
    prepare = preprocessing.MinMaxScaler() #MaxAbsScaler()
    x = prepare.fit_transform(mix.astype(float))
    histogrammes_train = x[0:histogrammes_train.shape[0], :]
    histogrammes_test = x[histogrammes_train.shape[0]:, :]
    exel_line=3
    nom_file_excel = 'knn_' + train_path[:-6] + '.xlsx'

    Parallel(n_jobs=num_cores)(delayed(fonctions.Reconnaisance)(3,histogrammes_train ,histogrammes_test,k_nn,label_train,label_test,exel_line,nom_file_excel) for k_nn in range(1, 9, 2))#convertir  histogrammes en 2 dimensions puisque chaque coeur calculer une matrice dépendante

    '''for k_nn in range(1, 9, 2):
        fonctions.Reconnaisance(1,histogrammes_train, histogrammes_test, k_nn,label_train, label_test,exel_line,nom_file_excel)'''
